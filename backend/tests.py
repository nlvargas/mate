from optimization import *
from collections import defaultdict
import pulp
from collections import OrderedDict, defaultdict
from openpyxl import load_workbook, Workbook
import json
import os

class Student:
    def __init__(self, id, attributes, preferences, modelAttributes, modelPreferences,disponibilities):
        self.id = id
        self.attributes = attributes
        self.preferences = preferences
        self.modelAttributes = modelAttributes
        self.modelPreferences = modelPreferences
        self.answered = True if preferences != ['None']*len(preferences) else False
        self.disponibilities = disponibilities
        self.priority = defaultdict(lambda: 100)
        self.flexibility = sum(int(i) for i in disponibilities)
        self.a = {}

    def __repr__(self):
        return "{}: Atributos: {} - Preferencias: {} - Disponibilidad: {}"\
               .format(self.id, self.attributes, self.preferences, self.a)

    def __str__(self):
        return self.__repr__()

    def get_disponibility(self, disponibilities_name):
        for i, j in zip(disponibilities_name, self.disponibilities):
            if j == 1:
                self.flexibility += 1
            if j != "#N/A":
                self.a[i] = int(j)
            else:
                self.a[i] = 1

    def get_priorities(self, groups):
        for i, preference in enumerate(self.preferences):
            groups_prefered = list(filter(lambda g: "Grupo {} - ".format(preference) in g, groups))
            for g in groups_prefered:
                self.priority[g] = i + 1

def count_options(students, options):
    attributes = {}
    for id in students:
        for a in students[id].attributes:
            if a not in attributes:
                attributes[a] = 1
            else:
                attributes[a] += 1
    opts = {attr: {} for attr in options}
    for attr in options:
        for opt in options[attr]:
            if opt != "None":
                opts[attr][opt] = attributes[opt]
    return opts

def generate_modules(params):
    capacity = params["capacity"]
    modules = params["modules"]
    attributes = params["student_attr"]
    students = params["students"]
    A = params["A"]
    I = list(students.keys()) 
    asignation = {mod: [s for s in students if int(students[s]["a"][mod]) == 1 and 
                        sum(int(students[s]["a"][m]) for m in modules) == 1] 
                  for mod in modules}
    ready = {mod: False for mod in modules}
    for i in I:
        isIn = False
        attrs = [students[i]["attributes"][attr] for attr in students[i]["attributes"]]
        available_modules = [mod for mod in modules if not ready[mod]]
        counter = {mod: {a: 0 for a in attrs} for mod in available_modules}
        total_counter = {mod: 0 for mod in available_modules}
        for mod in available_modules:
            for s in asignation[mod]:
                if i == s:
                    isIn = True
                    break
                for a in attrs:
                    if int(attributes[s][a]) == 1:
                        counter[mod][a] += 1
                        total_counter[mod] += 1
        if not isIn:
            # deberia ir ponderado por la cantidad de opciones de cada atributo
            avg_counter = {mod: {a: round(counter[mod][a]/len(asignation[mod]), 5) for a in attrs} 
                           for mod in modules}
            avg_total_counter = {mod: round(total_counter[mod]/len(asignation[mod]), 5) for mod in modules}
            min_mod = min(avg_total_counter, key=avg_total_counter.get)
            asignation[min_mod].append(i)
            if len(asignation[min_mod]) == capacity[min_mod]:
                ready[min_mod] = True
    # show results
    results = {mod: {a: 0 for a in A} for mod in modules}
    for mod in modules:
        for s in asignation[mod]:
            for a in A:
                if int(attributes[s][a]) == 1:
                    results[mod][a] += 1
    # print(results)
    for mod in modules:
        print(len(asignation[mod]))
    return asignation

def upload(request):
    f = request['file']
    wb = load_workbook(filename=f)
    attributes = request["attributes"]
    modules = request["modules"]
    preferences_number = int(request["preferencesNumber"])
    options = {attr: set() for attr in attributes}
    ws = wb["Alumnos"]
    students = {}
    parameters = ["student_fk"] + \
                 [a for a in attributes] + \
                 [d for d in modules] + \
                 [p for p in range(1, preferences_number + 1)]
    for row in ws.iter_rows(min_row=2, min_col=0, max_col=len(parameters)):
        p = OrderedDict()
        null_count = 0
        for i in range(0, len(parameters)):
            if i in range(1, len(attributes) + 1):
                options[parameters[i]].add(str(row[i].value))
            p[parameters[i]] = str(row[i].value)
            if row[i].value is None:
                null_count += 1
        if null_count >= len(parameters):
            break

        attr = {}
        for a in attributes:
            attr[a] = p[a]
        prefs = {}
        for i, pref in enumerate([p[pref] for pref in range(1, preferences_number + 1)]):
            prefs[i + 1] = pref
        s = Student(p["student_fk"],
                    [p[a] for a in attributes],
                    [p[pref] for pref in range(1, preferences_number + 1)],
                    attr,
                    prefs, 
                    [p[d] for d in modules])
        if modules:
            s.get_disponibility(modules)
        s.get_priorities(s.preferences)
        students[s.id] = s
    return count_options(students, options), [students[ob].__dict__ for ob in students]

def run_model(data):
    upper_number = data["minStudents"]
    attributes = {}
    for attr in data["options"]:
        attributes[attr] = {}
        for opt in data["options"][attr]:
            if opt in data["bounds"]:
                attributes[attr][opt] = data["bounds"][opt]
                if attributes[attr][opt]["min"] == "Min":
                    attributes[attr][opt]["min"] = 0
                if attributes[attr][opt]["max"] == "Max":
                    attributes[attr][opt]["max"] = upper_number
            else:
                attributes[attr][opt] = {"min": 0, 
                                         "max": upper_number, 
                                         "solo": False}
    options = [x for y in attributes for x in attributes[y]]

    for i in data["students"]:
        i["attributes"] = i["modelAttributes"]
        i["preferences"] = i["modelPreferences"]
        if i["preferences"][1] == "#N/A":
            i["answered"] = False
        else:
            i["answered"] = True
    student_attr = {}
    for i in data["students"]:
        student_attr[i["id"]] = {x: 0 for x in options}
        for j in i["attributes"]:
            student_attr[i["id"]][i["attributes"][j]] = 1
    students = {x['id']: x for x in data["students"]}

    preferences = {}
    for p in data['preferences']:
        if p in data['preferences']:
            preferences[p] = data['preferences'][p]
        else:
            preferences[p] = {"min": 0, "max": data['groupsNumber']}
    for s in students:
        a = {}
        for i, m in enumerate(data["modules"]):
            a[m] = students[s]["disponibilities"][i]
        students[s]["a"] = a
    params = {'attributes': attributes,
                'preferences': preferences,
                'groups_number': data['groupsNumber'],
                'lower_number': data['minStudents'],
                'upper_number': data['maxStudents'],
                'student_attr': student_attr,
                'students_preferences_number': data["preferencesNumber"],
                'students': students,
                'capacity': data['capacity'],
                'modules': data['modules'],
                'A': options,
                'tmax': int(data['tmax']),
                'sameDay': data['sameDay'],
                'fixedDay': data['fixedDay']}
    if data['modules']:
        groups = generate_modules(params)
        avg_pref = {mod: {pref: 0 for pref in preferences} for mod in groups}
        for mod in groups:
            for s in groups[mod]:
                for place in range(1, data["preferencesNumber"] + 1):
                    pref = students[s]["preferences"][place]
                    if pref != "#N/A":
                        avg_pref[mod][pref] += place
        avg_pref = {mod: {pref: avg_pref[mod][pref] / len(groups[mod]) for pref in preferences} for mod in groups}
        
        print(avg_pref)

        if data['sameDay']:
            sameDayPrefs = {}
            for pref in preferences:
                max_mod = None
                max_avg = None
                for mod in modules:
                    if max_mod is None or max_avg < avg_pref[mod][pref]:
                        max_avg = avg_pref[mod][pref]
                        max_mod = mod
                sameDayPrefs[pref] = max_mod

        module_preferences = {mod: preferences for mod in modules}
        module_students = {mod: {s: students[s] for s in groups[mod]} for mod in modules}
        module_preferences = {mod: round(len(groups[mod])/data['maxStudents']) for mod in modules}

        for mod in groups:
            for pref in preferences:
                if preferences[pref]["min"] != 0:
                    min_val = preferences[pref]["min"]
                    ratio = float(avg_pref[mod][pref]) / float(sum(avg_pref[module][pref] for module in modules)) 
                    module_preferences[mod][pref]["min"] = round(min_val * ratio)
                if preferences[pref]["max"] != data['groupsNumber']:
                    max_val = preferences[pref]["max"]
                    ratio = float(avg_pref[mod][pref]) / float(sum(avg_pref[module][pref] for module in modules)) 
                    module_preferences[mod][pref]["max"] = round(max_val * ratio)
                if data['sameDay']:
                    if sameDayPrefs[pref] != mod:
                        module_preferences[mod][pref]["min"] = 0
                        module_preferences[mod][pref]["max"] = 0
                    else:
                        module_preferences[mod][pref]["min"] =  preferences[pref]["min"]
                        module_preferences[mod][pref]["max"] =  preferences[pref]["max"]
                if data['fixedDay'][mod][pref]:
                    for m_ in modules:
                        if m_ != mod:
                           module_preferences[m_][pref]["min"] = 0
                           module_preferences[m_][pref]["max"] = 0 
            
            # res = json.dumps(run(params))
            # res = json.dumps(run_modules(params))
    else:
        res = json.dumps(run(params))


BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
filename = BASE_DIR + "/backend/data/SUS1000.xlsx"
attributes = {'Genero': {'Prefiero no decirlo': {'min': 0, 'max': 9, 'solo': False}, 'Hombre': {'min': 0, 'max': 10, 'solo': True}, 'Mujer': {'min': 0, 'max': 9, 'solo': False}}, 
              'Facultad': {'Medicina': {'min': 0, 'max': 9, 'solo': False}, 
                           'Letras': {'min': 0, 'max': 9, 'solo': False}, 
                           'Física': {'min': 0, 'max': 9, 'solo': False}, 
                           '#N/A': {'min': 0, 'max': 9, 'solo': False}, 
                           'Artes': {'min': 0, 'max': 9, 'solo': False}, 
                           'Agronomía e Ingeniería Forestal': {'min': 0, 'max': 9, 'solo': False}, 
                           'Ingeniería': {'min': 0, 'max': 9, 'solo': False}, 
                           'College': {'min': 0, 'max': 9, 'solo': False}, 
                           'Filosofía': {'min': 0, 'max': 9, 'solo': False}, 
                           'Ciencias Biológicas': {'min': 0, 'max': 9, 'solo': False}, 
                           'Arquitectura, Diseño y Estudios Urbanos': {'min': 0, 'max': 9, 'solo': False}, 
                           'Historia, Geografía y Ciencia Política': {'min': 0, 'max': 9, 'solo': False}, 
                           'Derecho': {'min': 0, 'max': 9, 'solo': False}, 
                           'Ciencias Económicas y Administrativas': {'min': 0, 'max': 9, 'solo': False}, 
                           'Comunicaciones': {'min': 0, 'max': 9, 'solo': False}, 
                           'Química': {'min': 0, 'max': 9, 'solo': False}, 
                           'Matemáticas': {'min': 0, 'max': 9, 'solo': False}, 
                           'Ciencias Sociales': {'min': 0, 'max': 9, 'solo': False}, 
                           'Educación': {'min': 0, 'max': 9, 'solo': False}}}
modules = ["Martes", "Miercoles"]
preferences = {'1': {'min': 0, 'max': 17}, 
               '2': {'min': 0, 'max': 17}, 
               '3': {'min': 0, 'max': 17}, 
               '4': {'min': 5, 'max': 9}, 
               '5': {'min': 0, 'max': 17}, 
               '6': {'min': 0, 'max': 17}}
preferencesNumber = 4
groupsNumber = 17
minStudents = 9
maxStudents = 10
request1 = {"attributes": attributes, "modules": modules, 
            "preferencesNumber": preferencesNumber, 
            "file": filename}

a, students = upload(request1)

options = {'Genero': {'Hombre': 68, 'Mujer': 99, 'Prefiero no decirlo': 1}, 
           'Facultad': {'Ciencias Sociales': 6, 'Química': 3, 'Historia, Geografía y Ciencia Política': 1, 'Comunicaciones': 2, 'Arquitectura, Diseño y Estudios Urbanos': 15, 'Artes': 1, 'Filosofía': 1, 'Letras': 2, 'Derecho': 8, '#N/A': 11, 'Medicina': 11, 'Agronomía e Ingeniería Forestal': 33, 'College': 20, 'Física': 1, 'Ciencias Biológicas': 8, 'Ingeniería': 25, 'Ciencias Económicas y Administrativas': 12, 'Educación': 6, 'Matemáticas': 2}
          }

fixedDay = {'Martes': {'1': False, '2': False, '3': True, '4': False, '5': False, '6': False}, 
            'Miercoles': {'1': False, '2': False, '3': False, '4': False, '5': False, '6': False}}

request2 = {"groupsNumber": groupsNumber,
            "minStudents": minStudents,
            "maxStudents": maxStudents,
            "preferencesNumber": preferencesNumber,
            "preferences": preferences,
            "tmax": 90,
            "sameDay": True,
            "options": options,
            "fixedDay": fixedDay,
            "bounds": {},
            'prefsBounds': {},
            "capacity": {m: len(students) for m in modules},
            "modules": modules,
            "attributes": a, 
            "students": students}

run_model(request2)