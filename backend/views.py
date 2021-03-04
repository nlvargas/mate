from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import api_view
from .optimization import run, run_modules, generate_modules
import openpyxl
from io import BytesIO
from collections import OrderedDict, defaultdict
from openpyxl import load_workbook, Workbook
import json


def count_options(students, options):
    attributes = {}
    for id in students:
        for a in students[id].attributes:
            if a not in attributes:
                attributes[a] = 1
            else:
                attributes[a] += 1
    opts = {attr: {} for attr in options}
    for attr in options:
        for opt in options[attr]:
            if opt != "None":
                opts[attr][opt] = attributes[opt]
    return opts

class Student:
    def __init__(self, id, attributes, preferences, modelAttributes, modelPreferences,disponibilities):
        self.id = id
        self.attributes = attributes
        self.preferences = preferences
        self.modelAttributes = modelAttributes
        self.modelPreferences = modelPreferences
        self.answered = True if preferences != ['None']*len(preferences) else False
        self.disponibilities = disponibilities
        self.priority = defaultdict(lambda: 100)
        self.flexibility = 1
        self.a = {}

    def __repr__(self):
        return "{}: Atributos: {} - Preferencias: {} - Disponibilidad: {}"\
               .format(self.id, self.attributes, self.preferences, self.a)

    def __str__(self):
        return self.__repr__()

    def get_disponibility(self, disponibilities_name):
        for i, j in zip(disponibilities_name, self.disponibilities):
            if j == 1:
                self.flexibility += 1
            if j != "#N/A":
                self.a[i] = int(j)
            else:
                self.a[i] = 1

    def get_priorities(self, groups):
        for i, preference in enumerate(self.preferences):
            groups_prefered = list(filter(lambda g: "Grupo {} - ".format(preference) in g, groups))
            for g in groups_prefered:
                self.priority[g] = i + 1

@api_view(['POST'])
def upload(request):
    f = request.FILES['file']
    wb = load_workbook(filename=BytesIO(f.read()))
    attributes = request.POST["attributes"].split(",")
    modules = request.POST["modules"].split(",")
    preferences_number = int(request.POST["preferencesNumber"][0])
    options = {attr: set() for attr in attributes}
    ws = wb["Alumnos"]
    students = {}
    parameters = ["student_fk"] + \
                 [a for a in attributes] + \
                 [d for d in modules] + \
                 [p for p in range(1, preferences_number + 1)]
    for row in ws.iter_rows(min_row=2, min_col=0, max_col=len(parameters)):
        p = OrderedDict()
        null_count = 0
        for i in range(0, len(parameters)):
            if i in range(1, len(attributes) + 1):
                options[parameters[i]].add(str(row[i].value))
            p[parameters[i]] = str(row[i].value)
            if row[i].value is None:
                null_count += 1
        if null_count >= len(parameters):
            break

        attr = {}
        for a in attributes:
            attr[a] = p[a]
        prefs = {}
        for i, pref in enumerate([p[pref] for pref in range(1, preferences_number + 1)]):
            prefs[i + 1] = pref
        s = Student(p["student_fk"],
                    [p[a] for a in attributes],
                    [p[pref] for pref in range(1, preferences_number + 1)],
                    attr,
                    prefs, 
                    [p[d] for d in modules])
        if modules:
            s.get_disponibility(modules)
        s.get_priorities(s.preferences)
        students[s.id] = s
    response = {"a": json.dumps(count_options(students, options)),
                "students": json.dumps([students[ob].__dict__ for ob in students])}
    return Response(response, status=status.HTTP_200_OK)


@api_view(['POST'])
def run_model(request):
    if request.method == 'POST':
        data = [json.loads(i) for i in request.POST.dict()][0]
        upper_number = data["minStudents"]
        attributes = {}
        for attr in data["options"]:
            attributes[attr] = {}
            for opt in data["options"][attr]:
                if opt in data["bounds"]:
                    attributes[attr][opt] = data["bounds"][opt]
                    if attributes[attr][opt]["min"] == "Min":
                        attributes[attr][opt]["min"] = 0
                    if attributes[attr][opt]["max"] == "Max":
                        attributes[attr][opt]["max"] = upper_number
                else:
                    attributes[attr][opt] = {"min": 0, 
                                             "max": upper_number, 
                                             "solo": False}
        options = [x for y in attributes for x in attributes[y]]


        for i in data["students"]:
            print(i)
            i["attributes"] = i["modelAttributes"]
            i["preferences"] = i["modelPreferences"]
            print(i["preferences"])
            if i["preferences"]["1"] == "#N/A":
                i["answered"] = False
            else:
                i["answered"] = True
        student_attr = {}
        for i in data["students"]:
            student_attr[i["id"]] = {x: 0 for x in options}
            for j in i["attributes"]:
                student_attr[i["id"]][i["attributes"][j]] = 1
        students = {x['id']: x for x in data["students"]}

        preferences = {}
        for p in data['preferences']:
            if p in data['prefsBounds']:
                preferences[p] = data['prefsBounds'][p]
            else:
                preferences[p] = {"min": 0, "max": data['groupsNumber']}
        for s in students:
            a = {}
            for i, m in enumerate(data["modules"]):
                a[m] = students[s]["disponibilities"][i]
            students[s]["a"] = a
        params = {'attributes': attributes,
                  'preferences': preferences,
                  'groups_number': data['groupsNumber'],
                  'lower_number': data['minStudents'],
                  'upper_number': data['maxStudents'],
                  'student_attr': student_attr,
                  'students_preferences_number': data["preferencesNumber"],
                  'students': students,
                  'capacity': data['capacity'],
                  'modules': data['modules'],
                  'A': options,
                  'tmax': int(data['tmax']),
                  'sameDay': data['sameDay'],
                  'fixedDay': data['fixedDay']}

        if data['modules']:
            groups = generate_modules(params)
            res = json.dumps(run_modules(params))
        else:
            res = json.dumps(run(params))
        return Response(res, status=status.HTTP_200_OK)

